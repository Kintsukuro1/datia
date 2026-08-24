import sqlite3
from typing import List
import openpyxl
from .base import sanitize_identifier, infer_sqlite_type, clean_cell_value

def import_excel_to_sqlite(excel_path: str, target_sqlite_path: str) -> List[str]:
    """
    Imports an Excel workbook (.xlsx, .xlsm, .xltx) into a SQLite database.
    Creates a dedicated table for each non-empty worksheet.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    created_tables = []
    conn = sqlite3.connect(target_sqlite_path)
    try:
        cur = conn.cursor()
        seen_table_names = set()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw_rows = list(ws.iter_rows(values_only=True))
            if not raw_rows:
                continue

            raw_headers = None
            data_rows = []
            for row in raw_rows:
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue
                if raw_headers is None:
                    raw_headers = row
                else:
                    data_rows.append(row)

            if not raw_headers:
                continue

            clean_tbl = sanitize_identifier(sheet_name, fallback_prefix="hoja")
            final_tbl = clean_tbl
            counter = 1
            while final_tbl in seen_table_names:
                final_tbl = f"{clean_tbl}_{counter}"
                counter += 1
            seen_table_names.add(final_tbl)

            headers = []
            seen_cols = set()
            for idx, h in enumerate(raw_headers):
                clean_col = sanitize_identifier(str(h) if h is not None else "", fallback_prefix=f"col_{idx+1}")
                col_final = clean_col
                c_count = 1
                while col_final in seen_cols:
                    col_final = f"{clean_col}_{c_count}"
                    c_count += 1
                seen_cols.add(col_final)
                headers.append(col_final)

            sample_rows = data_rows[:100]
            col_types = []
            for c_idx in range(len(headers)):
                col_samples = [r[c_idx] for r in sample_rows if c_idx < len(r)]
                col_types.append(infer_sqlite_type(col_samples))

            cols_def = ", ".join([f'"{h}" {t}' for h, t in zip(headers, col_types)])
            cur.execute(f'DROP TABLE IF EXISTS "{final_tbl}";')
            cur.execute(f'CREATE TABLE "{final_tbl}" ({cols_def});')

            placeholders = ", ".join(["?"] * len(headers))
            insert_sql = f'INSERT INTO "{final_tbl}" VALUES ({placeholders});'

            cleaned_data = []
            for r in data_rows:
                row_vals = []
                for c_idx in range(len(headers)):
                    val = r[c_idx] if c_idx < len(r) else None
                    row_vals.append(clean_cell_value(val))
                cleaned_data.append(row_vals)

            cur.executemany(insert_sql, cleaned_data)
            created_tables.append(final_tbl)

        conn.commit()
    finally:
        conn.close()
        wb.close()

    if not created_tables:
        raise ValueError("El archivo Excel no contiene hojas de cálculo con datos tabulares válidos.")

    return created_tables
