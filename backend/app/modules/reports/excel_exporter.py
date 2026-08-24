import io
import datetime
from typing import Union
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.modules.admin_catalog.schemas import ReportExportData, ReportExportRequest
from app.modules.reports.data_compiler import ReportDataCompiler

class ExcelExporter:
    """
    Generates structured Excel workbooks with executive summaries and data tables using openpyxl.
    """

    @classmethod
    def generate_excel(cls, data: Union[ReportExportData, ReportExportRequest]) -> bytes:
        wb = Workbook()

        # Sheet 1: Resumen Ejecutivo
        ws_summary = wb.active
        ws_summary.title = "Resumen Ejecutivo"
        ws_summary.views.sheetView[0].showGridLines = True

        font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_section = Font(name="Calibri", size=12, bold=True, color="312E81")
        font_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
        font_normal = Font(name="Calibri", size=10, color="334155")
        font_kpi_val = Font(name="Calibri", size=14, bold=True, color="4F46E5")

        fill_header = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        thin_side = Side(style='thin', color="CBD5E1")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        ws_summary.merge_cells("A1:F2")
        title_cell = ws_summary["A1"]
        title_cell.value = "DATIA - INFORME EJECUTIVO DE NEGOCIO"
        title_cell.font = font_title
        title_cell.fill = fill_header
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_summary["A4"] = "Consulta Analizada:"
        ws_summary["A4"].font = font_bold
        ws_summary["B4"] = data.question
        ws_summary["B4"].font = font_normal

        ws_summary["A5"] = "Fecha Generación:"
        ws_summary["A5"].font = font_bold
        ws_summary["B5"] = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        ws_summary["B5"].font = font_normal

        ws_summary["A6"] = "Base de Datos:"
        ws_summary["A6"].font = font_bold
        ws_summary["B6"] = data.target_database or "demo_corporativa.db (SQLite)"
        ws_summary["B6"].font = font_normal

        current_row = 8

        if data.kpis:
            ws_summary.cell(row=current_row, column=1, value="INDICADORES CLAVE (KPIs)").font = font_section
            current_row += 1
            for k in data.kpis:
                ws_summary.cell(row=current_row, column=1, value=k.title).font = font_bold
                val_cell = ws_summary.cell(row=current_row, column=2, value=k.value)
                val_cell.font = font_kpi_val
                if k.subtitle:
                    ws_summary.cell(row=current_row, column=3, value=k.subtitle).font = font_normal
                current_row += 1
            current_row += 1

        ws_summary.cell(row=current_row, column=1, value="1. DIAGNÓSTICO & IMPACTO EN EL NEGOCIO").font = font_section
        current_row += 1
        overview = ReportDataCompiler.compile_summary_text(data)
        ws_summary.cell(row=current_row, column=1, value=overview or "").font = font_normal
        current_row += 1

        if data.executive_report and data.executive_report.business_impact:
            ws_summary.cell(row=current_row, column=1, value=f"Impacto: {data.executive_report.business_impact}").font = font_normal
            current_row += 1
        current_row += 1

        findings = ReportDataCompiler.compile_findings(data)
        if findings:
            ws_summary.cell(row=current_row, column=1, value="2. HALLAZGOS CLAVE").font = font_section
            current_row += 1
            for f in findings:
                ws_summary.cell(row=current_row, column=1, value=f"• {f}").font = font_normal
                current_row += 1
            current_row += 1

        recs = ReportDataCompiler.compile_recommendations(data)
        if recs:
            ws_summary.cell(row=current_row, column=1, value="3. RECOMENDACIONES ESTRATÉGICAS").font = font_section
            current_row += 1
            for i, r in enumerate(recs, 1):
                ws_summary.cell(row=current_row, column=1, value=f"{i}. {r}").font = font_normal
                current_row += 1
            current_row += 1

        ws_summary.cell(row=current_row, column=1, value="4. TRAZABILIDAD TÉCNICA").font = font_section
        current_row += 1
        trace = data.traceability
        ws_summary.cell(row=current_row, column=1, value=f"Estado AST: {trace.validation_status if trace else 'APROBADO'} | Filas: {trace.rows_returned if trace else len(data.data_rows)} | Latencia: {trace.execution_time_ms if trace else 0} ms").font = font_normal
        current_row += 1
        if trace and trace.sql_executed:
            ws_summary.cell(row=current_row, column=1, value="SQL Ejecutado:").font = font_bold
            ws_summary.cell(row=current_row + 1, column=1, value=trace.sql_executed).font = Font(name="Consolas", size=9, color="312E81")

        ws_summary.column_dimensions['A'].width = 28
        ws_summary.column_dimensions['B'].width = 60
        ws_summary.column_dimensions['C'].width = 30

        # Sheet 2: Datos
        ws_data = wb.create_sheet(title="Datos")
        ws_data.views.sheetView[0].showGridLines = True

        columns = ReportDataCompiler.compile_columns(data)

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all

        ws_data.row_dimensions[1].height = 24

        for row_idx, row_dict in enumerate(data.data_rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                val = row_dict.get(col_name)
                cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_normal
                cell.border = border_all
                if row_idx % 2 == 1:
                    cell.fill = fill_zebra
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        for col in ws_data.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws_data.column_dimensions[col_letter].width = max(12, min(max_len + 4, 50))

        if columns and data.data_rows:
            last_col_letter = get_column_letter(len(columns))
            ws_data.auto_filter.ref = f"A1:{last_col_letter}{len(data.data_rows) + 1}"

        out_buffer = io.BytesIO()
        wb.save(out_buffer)
        excel_bytes = out_buffer.getvalue()
        out_buffer.close()
        return excel_bytes
