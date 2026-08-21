import io
import base64
import datetime
from typing import Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepTogether, HRFlowable
)

from app.schemas.report_schema import ReportExportRequest

class NumberedCanvas:
    """Helper for page numbering on footer."""
    pass

class ReportGeneratorService:
    """
    Generates professional, presentation-ready PDF reports and structured Excel (.xlsx) workbooks
    from pre-computed QueryResult data.
    """

    @classmethod
    def generate_pdf(cls, data: ReportExportRequest) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()

        # Custom Palette
        COLOR_PRIMARY = colors.HexColor("#312E81")     # Indigo dark
        COLOR_ACCENT = colors.HexColor("#4F46E5")      # Indigo bright
        COLOR_EMERALD = colors.HexColor("#059669")     # Emerald green
        COLOR_AMBER = colors.HexColor("#D97706")       # Amber
        COLOR_DARK = colors.HexColor("#0F172A")        # Dark slate
        COLOR_MUTED = colors.HexColor("#64748B")       # Slate muted
        COLOR_BG_LIGHT = colors.HexColor("#F8FAFC")    # Very light gray
        COLOR_BORDER = colors.HexColor("#E2E8F0")      # Light border

        # Custom Paragraph Styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            textColor=COLOR_PRIMARY,
            spaceAfter=4
        )

        subtitle_style = ParagraphStyle(
            'DocSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=12,
            textColor=COLOR_MUTED,
            spaceAfter=10
        )

        section_heading = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=11,
            leading=14,
            textColor=COLOR_ACCENT,
            spaceBefore=8,
            spaceAfter=4
        )

        body_style = ParagraphStyle(
            'BodyDark',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=13,
            textColor=COLOR_DARK
        )

        body_bold = ParagraphStyle(
            'BodyBold',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=13,
            textColor=COLOR_DARK
        )

        sql_code_style = ParagraphStyle(
            'SqlCode',
            parent=styles['Code'],
            fontName='Courier',
            fontSize=8,
            leading=11,
            textColor=COLOR_PRIMARY
        )

        story = []

        # 1. Header Banner
        header_table = Table(
            [
                [
                    Paragraph("<b>DATIA</b> | Executive Analytics", title_style),
                    Paragraph(f"<b>Fecha:</b> {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}<br/><b>BD:</b> {data.target_database or 'SQLite Demo'}", subtitle_style)
                ]
            ],
            colWidths=[360, 180]
        )
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        story.append(header_table)
        story.append(HRFlowable(width="100%", thickness=1.5, color=COLOR_ACCENT, spaceBefore=4, spaceAfter=8))

        # 2. Question / Prompt Box
        question_p = Paragraph(f"<b>Consulta Analizada:</b> <i>\"{data.question}\"</i>", body_style)
        q_table = Table([[question_p]], colWidths=[540])
        q_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), COLOR_BG_LIGHT),
            ('BOX', (0, 0), (-1, -1), 0.75, COLOR_BORDER),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(q_table)
        story.append(Spacer(1, 8))

        # 3. KPI Cards Grid (if available)
        if data.kpis:
            kpi_cols = min(len(data.kpis), 4)
            col_w = 540 / kpi_cols
            kpi_cells = []
            for k in data.kpis[:kpi_cols]:
                cell_content = [
                    Paragraph(f"<font size=7 color='#64748B'><b>{k.title.upper()}</b></font>", body_style),
                    Spacer(1, 2),
                    Paragraph(f"<font size=13 color='#312E81'><b>{k.value}</b></font>", body_style)
                ]
                if k.subtitle:
                    cell_content.append(Paragraph(f"<font size=7 color='#94A3B8'>{k.subtitle}</font>", body_style))
                kpi_cells.append(cell_content)

            kpi_table = Table([kpi_cells], colWidths=[col_w] * kpi_cols)
            kpi_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), COLOR_BG_LIGHT),
                ('BOX', (0, 0), (-1, -1), 0.5, COLOR_BORDER),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, COLOR_BORDER),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(kpi_table)
            story.append(Spacer(1, 8))

        # 4. Diagnóstico & Contexto General
        story.append(Paragraph("1. Diagnóstico & Contexto General", section_heading))
        overview_text = ""
        if data.executive_report and data.executive_report.overview:
            overview_text = data.executive_report.overview
        elif data.summary_text:
            overview_text = data.summary_text
        else:
            overview_text = "Se procesó la consulta satisfactoriamente sobre el esquema corporativo autorizado."

        risk_badge = ""
        if data.executive_report and data.executive_report.risk_level:
            risk_badge = f" [Nivel de Riesgo: <b>{data.executive_report.risk_level}</b>]"

        diag_content = [Paragraph(f"{overview_text}{risk_badge}", body_style)]
        if data.executive_report and data.executive_report.business_impact:
            diag_content.append(Spacer(1, 4))
            diag_content.append(Paragraph(f"<b>Impacto en el Negocio:</b> {data.executive_report.business_impact}", body_style))

        diag_table = Table([[diag_content]], colWidths=[540])
        diag_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), COLOR_BG_LIGHT),
            ('BOX', (0, 0), (-1, -1), 0.5, COLOR_BORDER),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(diag_table)
        story.append(Spacer(1, 8))

        # 5. Hallazgos Clave
        findings = data.executive_report.key_findings if (data.executive_report and data.executive_report.key_findings) else []
        if not findings and data.data_rows:
            findings = [
                f"Se procesaron {len(data.data_rows)} registros de la base de datos activa.",
                f"Columnas analizadas: {', '.join(data.data_columns[:5])}."
            ]

        if findings:
            story.append(Paragraph("2. Hallazgos Clave & Puntos Críticos", section_heading))
            findings_data = []
            for f in findings:
                p = Paragraph(f"• {f}", body_style)
                findings_data.append([p])

            findings_table = Table(findings_data, colWidths=[540])
            findings_table.setStyle(TableStyle([
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(findings_table)
            story.append(Spacer(1, 6))

        # 6. Recomendaciones Estratégicas
        recommendations = data.executive_report.recommendations if (data.executive_report and data.executive_report.recommendations) else []
        if recommendations:
            story.append(Paragraph("3. Recomendaciones Estratégicas Accionables", section_heading))
            recs_data = []
            for i, r in enumerate(recommendations, 1):
                p = Paragraph(f"<b>{i}.</b> {r}", body_style)
                recs_data.append([p])

            recs_table = Table(recs_data, colWidths=[540])
            recs_table.setStyle(TableStyle([
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(recs_table)
            story.append(Spacer(1, 6))

        # 7. Chart Image (if present)
        if data.chart_image_base64:
            try:
                img_data_str = data.chart_image_base64
                if "," in img_data_str:
                    img_data_str = img_data_str.split(",", 1)[1]
                img_bytes = base64.b64decode(img_data_str)
                img_stream = io.BytesIO(img_bytes)

                chart_img = Image(img_stream, width=500, height=200)
                chart_img.hAlign = 'CENTER'

                chart_block = [
                    Paragraph("4. Visualización Analítica", section_heading),
                    chart_img,
                    Spacer(1, 6)
                ]
                story.append(KeepTogether(chart_block))
            except Exception:
                pass

        # 8. Technical Traceability & Governance Block
        trace = data.traceability
        trace_title = "4. Trazabilidad Técnica & Gobernanza" if not data.chart_image_base64 else "5. Trazabilidad Técnica & Gobernanza"
        story.append(Paragraph(trace_title, section_heading))

        val_status = trace.validation_status if trace else "APROBADO"
        rows_ret = trace.rows_returned if trace else len(data.data_rows)
        exec_ms = trace.execution_time_ms if trace else 0

        trace_meta_p = Paragraph(
            f"<b>Estado AST:</b> {val_status} | <b>Filas:</b> {rows_ret} | <b>Latencia:</b> {exec_ms} ms",
            body_style
        )

        trace_cells = [[trace_meta_p]]
        if trace and trace.sql_executed:
            sql_p = Paragraph(f"<b>SQL Validado:</b><br/>{trace.sql_executed}", sql_code_style)
            trace_cells.append([sql_p])

        trace_table = Table(trace_cells, colWidths=[540])
        trace_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), COLOR_BG_LIGHT),
            ('BOX', (0, 0), (-1, -1), 0.5, COLOR_BORDER),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(trace_table)

        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    @classmethod
    def generate_excel(cls, data: ReportExportRequest) -> bytes:
        wb = Workbook()

        # Sheet 1: Resumen Ejecutivo
        ws_summary = wb.active
        ws_summary.title = "Resumen Ejecutivo"
        ws_summary.views.sheetView[0].showGridLines = True

        # Styles
        font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_section = Font(name="Calibri", size=12, bold=True, color="312E81")
        font_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
        font_normal = Font(name="Calibri", size=10, color="334155")
        font_kpi_val = Font(name="Calibri", size=14, bold=True, color="4F46E5")

        fill_header = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
        fill_subhead = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        thin_side = Side(style='thin', color="CBD5E1")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Title Banner
        ws_summary.merge_cells("A1:F2")
        title_cell = ws_summary["A1"]
        title_cell.value = "DATIA - INFORME EJECUTIVO DE NEGOCIO"
        title_cell.font = font_title
        title_cell.fill = fill_header
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Metadata
        ws_summary["A4"] = "Consulta Analizada:"
        ws_summary["A4"].font = font_bold
        ws_summary["B4"] = data.question
        ws_summary["B4"].font = font_normal

        ws_summary["A5"] = "Fecha Generación:"
        ws_summary["A5"].font = font_bold
        ws_summary["B5"] = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        ws_summary["B5"].font = font_normal

        ws_summary["A6"] = "Base de Datos:"
        ws_summary["A6"].font = font_bold
        ws_summary["B6"] = data.target_database or "demo_corporativa.db (SQLite)"
        ws_summary["B6"].font = font_normal

        current_row = 8

        # KPIs Section
        if data.kpis:
            ws_summary.cell(row=current_row, column=1, value="INDICADORES CLAVE (KPIs)").font = font_section
            current_row += 1
            for k in data.kpis:
                ws_summary.cell(row=current_row, column=1, value=k.title).font = font_bold
                val_cell = ws_summary.cell(row=current_row, column=2, value=k.value)
                val_cell.font = font_kpi_val
                if k.subtitle:
                    ws_summary.cell(row=current_row, column=3, value=k.subtitle).font = font_normal
                current_row += 1
            current_row += 1

        # Diagnóstico Section
        ws_summary.cell(row=current_row, column=1, value="1. DIAGNÓSTICO & IMPACTO EN EL NEGOCIO").font = font_section
        current_row += 1
        overview = data.executive_report.overview if (data.executive_report and data.executive_report.overview) else data.summary_text
        ws_summary.cell(row=current_row, column=1, value=overview or "").font = font_normal
        current_row += 1

        if data.executive_report and data.executive_report.business_impact:
            ws_summary.cell(row=current_row, column=1, value=f"Impacto: {data.executive_report.business_impact}").font = font_normal
            current_row += 1
        current_row += 1

        # Hallazgos Clave
        findings = data.executive_report.key_findings if (data.executive_report and data.executive_report.key_findings) else []
        if findings:
            ws_summary.cell(row=current_row, column=1, value="2. HALLAZGOS CLAVE").font = font_section
            current_row += 1
            for f in findings:
                ws_summary.cell(row=current_row, column=1, value=f"• {f}").font = font_normal
                current_row += 1
            current_row += 1

        # Recomendaciones
        recs = data.executive_report.recommendations if (data.executive_report and data.executive_report.recommendations) else []
        if recs:
            ws_summary.cell(row=current_row, column=1, value="3. RECOMENDACIONES ESTRATÉGICAS").font = font_section
            current_row += 1
            for i, r in enumerate(recs, 1):
                ws_summary.cell(row=current_row, column=1, value=f"{i}. {r}").font = font_normal
                current_row += 1
            current_row += 1

        # Trazabilidad
        ws_summary.cell(row=current_row, column=1, value="4. TRAZABILIDAD TÉCNICA").font = font_section
        current_row += 1
        trace = data.traceability
        ws_summary.cell(row=current_row, column=1, value=f"Estado AST: {trace.validation_status if trace else 'APROBADO'} | Filas: {trace.rows_returned if trace else len(data.data_rows)} | Latencia: {trace.execution_time_ms if trace else 0} ms").font = font_normal
        current_row += 1
        if trace and trace.sql_executed:
            ws_summary.cell(row=current_row, column=1, value="SQL Ejecutado:").font = font_bold
            ws_summary.cell(row=current_row + 1, column=1, value=trace.sql_executed).font = Font(name="Consolas", size=9, color="312E81")

        # Column widths for Summary
        ws_summary.column_dimensions['A'].width = 28
        ws_summary.column_dimensions['B'].width = 60
        ws_summary.column_dimensions['C'].width = 30

        # Sheet 2: Datos
        ws_data = wb.create_sheet(title="Datos")
        ws_data.views.sheetView[0].showGridLines = True

        columns = data.data_columns
        if not columns and data.data_rows:
            columns = list(data.data_rows[0].keys())

        # Header Row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all

        ws_data.row_dimensions[1].height = 24

        # Data Rows
        for row_idx, row_dict in enumerate(data.data_rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                val = row_dict.get(col_name)
                cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_normal
                cell.border = border_all
                if row_idx % 2 == 1:
                    cell.fill = fill_zebra
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # Auto-adjust column widths
        for col in ws_data.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws_data.column_dimensions[col_letter].width = max(12, min(max_len + 4, 50))

        # Enable AutoFilter on Data Sheet
        if columns and data.data_rows:
            last_col_letter = get_column_letter(len(columns))
            ws_data.auto_filter.ref = f"A1:{last_col_letter}{len(data.data_rows) + 1}"

        out_buffer = io.BytesIO()
        wb.save(out_buffer)
        excel_bytes = out_buffer.getvalue()
        out_buffer.close()
        return excel_bytes
