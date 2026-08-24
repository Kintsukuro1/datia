import os
import re
import csv
import sqlite3
import datetime
from typing import List, Tuple, Dict, Any, Optional
import openpyxl

def sanitize_identifier(name: str, fallback_prefix: str = "col") -> str:
    """
    Sanitizes table or column names to be valid SQLite identifiers.
    Replaces accents, spaces, and special characters with underscores.
    """
    if not name:
        return fallback_prefix
    
    clean = str(name).strip().lower()
    # Normalize accents
    clean = (
        clean.replace('á', 'a')
        .replace('é', 'e')
        .replace('í', 'i')
        .replace('ó', 'o')
        .replace('ú', 'u')
        .replace('ñ', 'n')
    )
    # Replace non-alphanumeric with underscores
    clean = re.sub(r'[^a-zA-Z0-9_]+', '_', clean)
    clean = re.sub(r'_+', '_', clean).strip('_')
    
    if not clean or clean[0].isdigit():
        clean = f"{fallback_prefix}_{clean}"
    return clean

def infer_sqlite_type(values: List[Any]) -> str:
    """
    Infers the most appropriate SQLite column type from a sample of non-null values.
    """
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_empty:
        return "TEXT"
    
    is_int = True
    is_float = True
    for v in non_empty:
        s = str(v).strip().replace("$", "").replace("€", "").replace("%", "")
        # Allow numbers with comma decimal if no period
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        try:
            int(s)
        except ValueError:
            is_int = False
            try:
                float(s)
            except ValueError:
                is_float = False
                break
    
    if is_int:
        return "INTEGER"
    if is_float:
        return "REAL"
    return "TEXT"

def clean_cell_value(val: Any) -> Any:
    """
    Cleans and formats cell values for SQLite insertion.
    """
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.isoformat()
    if isinstance(val, (int, float, bool)):
        return val
    s = str(val).strip()
    return s if s != "" else None

def import_csv_to_sqlite(csv_path: str, target_sqlite_path: str, table_name: Optional[str] = None) -> List[str]:
    """
    Imports a CSV file into a SQLite database with automatic delimiter detection,
    encoding recovery, and type inference.
    """
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252", "iso-8859-1"]
    content = None
    used_encoding = "utf-8"

    for enc in encodings:
        try:
            with open(csv_path, "r", encoding=enc) as f:
                content = f.read(65536)
                used_encoding = enc
                break
        except UnicodeDecodeError:
            continue

    if content is None:
        raise ValueError("No se pudo decodificar el archivo CSV con los juegos de caracteres estándar.")

    # Detect delimiter
    sample = content[:4096]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
        delimiter = dialect.delimiter
    except Exception:
        # Fallback heuristic
        counts = {d: sample.count(d) for d in [",", ";", "\t", "|"]}
        delimiter = max(counts, key=counts.get) if any(counts.values()) else ","

    # Parse full CSV
    with open(csv_path, "r", encoding=used_encoding, errors="replace") as f:
        reader = csv.reader(f, delimiter=delimiter)
        raw_headers = None
        data_rows = []
        for row in reader:
            if not row or all(str(cell).strip() == "" for cell in row):
                continue
            if raw_headers is None:
                raw_headers = row
            else:
                data_rows.append(row)

    if not raw_headers:
        raise ValueError("El archivo CSV está vacío o no contiene encabezados legibles.")

    # Sanitize and ensure unique column names
    headers = []
    seen_cols = set()
    for idx, h in enumerate(raw_headers):
        clean_col = sanitize_identifier(h, fallback_prefix=f"col_{idx+1}")
        col_final = clean_col
        c_count = 1
        while col_final in seen_cols:
            col_final = f"{clean_col}_{c_count}"
            c_count += 1
        seen_cols.add(col_final)
        headers.append(col_final)

    # Determine final table name
    if not table_name:
        base_file = os.path.splitext(os.path.basename(csv_path))[0]
        # Remove any leading internal temporary prefix like raw_1234567890_
        base_file = re.sub(r'^raw_\d+_', '', base_file)
        table_name = sanitize_identifier(base_file, fallback_prefix="tabla_csv")

    # Type inference per column
    col_types = []
    sample_rows = data_rows[:100]
    for c_idx in range(len(headers)):
        col_samples = [r[c_idx] for r in sample_rows if c_idx < len(r)]
        col_types.append(infer_sqlite_type(col_samples))

    # Create table in SQLite
    conn = sqlite3.connect(target_sqlite_path)
    cur = conn.cursor()

    cols_def = ", ".join([f'"{h}" {t}' for h, t in zip(headers, col_types)])
    cur.execute(f'DROP TABLE IF EXISTS "{table_name}";')
    cur.execute(f'CREATE TABLE "{table_name}" ({cols_def});')

    # Insert data
    placeholders = ", ".join(["?"] * len(headers))
    insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders});'

    cleaned_data = []
    for r in data_rows:
        row_vals = []
        for c_idx in range(len(headers)):
            val = r[c_idx] if c_idx < len(r) else None
            row_vals.append(clean_cell_value(val))
        cleaned_data.append(row_vals)

    cur.executemany(insert_sql, cleaned_data)
    conn.commit()
    conn.close()

    return [table_name]

def import_excel_to_sqlite(excel_path: str, target_sqlite_path: str) -> List[str]:
    """
    Imports an Excel workbook (.xlsx, .xlsm, .xltx) into a SQLite database.
    Creates a dedicated table for each non-empty worksheet.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    created_tables = []
    conn = sqlite3.connect(target_sqlite_path)
    cur = conn.cursor()

    seen_table_names = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            continue

        # Find first non-empty row for headers
        raw_headers = None
        data_rows = []
        for row in raw_rows:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            if raw_headers is None:
                raw_headers = row
            else:
                data_rows.append(row)

        if not raw_headers:
            continue

        # Sanitize table name
        clean_tbl = sanitize_identifier(sheet_name, fallback_prefix="hoja")
        final_tbl = clean_tbl
        counter = 1
        while final_tbl in seen_table_names:
            final_tbl = f"{clean_tbl}_{counter}"
            counter += 1
        seen_table_names.add(final_tbl)

        # Sanitize headers
        headers = []
        seen_cols = set()
        for idx, h in enumerate(raw_headers):
            clean_col = sanitize_identifier(str(h) if h is not None else "", fallback_prefix=f"col_{idx+1}")
            col_final = clean_col
            c_count = 1
            while col_final in seen_cols:
                col_final = f"{clean_col}_{c_count}"
                c_count += 1
            seen_cols.add(col_final)
            headers.append(col_final)

        # Infer types
        sample_rows = data_rows[:100]
        col_types = []
        for c_idx in range(len(headers)):
            col_samples = [r[c_idx] for r in sample_rows if c_idx < len(r)]
            col_types.append(infer_sqlite_type(col_samples))

        # Create table
        cols_def = ", ".join([f'"{h}" {t}' for h, t in zip(headers, col_types)])
        cur.execute(f'DROP TABLE IF EXISTS "{final_tbl}";')
        cur.execute(f'CREATE TABLE "{final_tbl}" ({cols_def});')

        # Insert data
        placeholders = ", ".join(["?"] * len(headers))
        insert_sql = f'INSERT INTO "{final_tbl}" VALUES ({placeholders});'

        cleaned_data = []
        for r in data_rows:
            row_vals = []
            for c_idx in range(len(headers)):
                val = r[c_idx] if c_idx < len(r) else None
                row_vals.append(clean_cell_value(val))
            cleaned_data.append(row_vals)

        cur.executemany(insert_sql, cleaned_data)
        created_tables.append(final_tbl)

    conn.commit()
    conn.close()
    wb.close()

    if not created_tables:
        raise ValueError("El archivo Excel no contiene hojas de cálculo con datos tabulares válidos.")

    return created_tables

def convert_uploaded_file_to_sqlite(source_path: str, ext: str, target_sqlite_path: str, table_name: Optional[str] = None) -> List[str]:
    """
    Orchestrates the conversion of any supported file format (.csv, .xlsx, .xls, .sql, .sqlite, .db)
    into a production-ready SQLite database and returns the list of detected table names.
    """
    clean_ext = ext.lower().strip()

    if clean_ext in [".csv", ".txt", ".tsv"]:
        return import_csv_to_sqlite(source_path, target_sqlite_path, table_name=table_name)
    
    elif clean_ext in [".xlsx", ".xlsm", ".xltx", ".xls"]:
        return import_excel_to_sqlite(source_path, target_sqlite_path)
    
    elif clean_ext == ".sql":
        with open(source_path, "r", encoding="utf-8", errors="ignore") as sql_file:
            sql_script = sql_file.read()
        conn = sqlite3.connect(target_sqlite_path)
        conn.executescript(sql_script)
        cur = conn.cursor()
        tables = [
            r[0] for r in cur.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()
            if not r[0].startswith("sqlite_")
        ]
        conn.close()
        return tables

    elif clean_ext in [".sqlite", ".db", ".sqlite3"]:
        # If source is already sqlite, check tables
        conn = sqlite3.connect(target_sqlite_path)
        cur = conn.cursor()
        tables = [
            r[0] for r in cur.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()
            if not r[0].startswith("sqlite_")
        ]
        conn.close()
        if not tables:
            raise ValueError("El archivo SQLite subido no contiene tablas válidas.")
        return tables

    else:
        raise ValueError(f"Formato no soportado: {clean_ext}. Formatos permitidos: .sqlite, .db, .sqlite3, .csv, .xlsx, .xls, .sql")
